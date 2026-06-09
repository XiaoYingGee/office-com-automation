"""
gen_fixture.py — Generate benchmark fixture workbooks at three size tiers.

Tiers:
  empty   1 sheet,  no data rows                          (baseline)
  medium  3 sheets x 1000 rows x 50 cols   = 150K cells   (all backends)
  large   10 sheets x 20000 rows x 100 cols = 20M cells   (persistent backends)

Content is seeded-random (reproducible): column 0 is a row label string, the
remaining columns are a deterministic mix of numbers and short strings.

Usage:
  python gen_fixture.py                      # generate all tiers (skip if present)
  python gen_fixture.py --size medium,large  # subset
  python gen_fixture.py --force              # regenerate even if present
  python gen_fixture.py --out-dir <path>     # default: ../../../benchmarks/fixtures
"""
import argparse
import os
import random
import sys
import time

try:
    from openpyxl import Workbook
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# tier -> (sheets, rows, cols)
TIERS = {
    "empty":  (1, 0, 0),
    "medium": (3, 1000, 50),
    "large":  (10, 20000, 100),
}

SEED = 20260609
WORDS = ["alpha", "bravo", "charlie", "delta", "echo", "foxtrot",
         "golf", "hotel", "india", "juliet", "kilo", "lima"]


def _cell_value(rng, r, c):
    """Deterministic mixed content: col 0 label, else number or occasional word."""
    if c == 0:
        return f"R{r}"
    # ~15% strings, rest numbers — exercises shared-strings + numeric paths
    if rng.random() < 0.15:
        return WORDS[rng.randrange(len(WORDS))]
    return round(rng.random() * 100000, 2)


def gen_tier(name, out_dir, force=False):
    sheets, rows, cols = TIERS[name]
    path = os.path.join(out_dir, f"bench_{name}.xlsx")

    if os.path.isfile(path) and not force:
        size_mb = os.path.getsize(path) / (1024 * 1024)
        print(f"  [{name}] exists, skip ({size_mb:.1f} MB) — use --force to regenerate")
        return path

    rng = random.Random(SEED)
    t0 = time.perf_counter()

    # write_only mode streams rows to disk — essential for the 20M-cell large tier.
    wb = Workbook(write_only=True)
    for si in range(sheets):
        ws = wb.create_sheet(title=f"Sheet{si + 1}")
        for r in range(1, rows + 1):
            ws.append([_cell_value(rng, r, c) for c in range(cols)])

    # write_only workbooks must have at least one sheet; empty tier gets a bare sheet.
    if sheets == 0:
        wb.create_sheet(title="Sheet1")

    os.makedirs(out_dir, exist_ok=True)
    wb.save(path)

    elapsed = time.perf_counter() - t0
    size_mb = os.path.getsize(path) / (1024 * 1024)
    print(f"  [{name}] {sheets} sheet(s) x {rows} x {cols} -> {size_mb:.1f} MB in {elapsed:.1f}s")
    return path


def main():
    parser = argparse.ArgumentParser(description="Generate benchmark fixtures")
    parser.add_argument("--size", default="empty,medium,large",
                        help="Comma-separated tiers (empty,medium,large)")
    parser.add_argument("--out-dir", default=None, help="Output directory")
    parser.add_argument("--force", action="store_true", help="Regenerate even if present")
    args = parser.parse_args()

    out_dir = args.out_dir or os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "..", "..", "benchmarks", "fixtures"
    )
    out_dir = os.path.abspath(out_dir)

    tiers = [t.strip() for t in args.size.split(",") if t.strip()]
    unknown = [t for t in tiers if t not in TIERS]
    if unknown:
        print(f"Unknown tier(s): {unknown}; valid: {list(TIERS)}", file=sys.stderr)
        sys.exit(1)

    print(f"Fixtures -> {out_dir}")
    for t in tiers:
        gen_tier(t, out_dir, force=args.force)


if __name__ == "__main__":
    main()
