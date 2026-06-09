"""Generate benchmark fixture files: empty, medium (~50MB), large (~100MB).

Uses openpyxl to create .xlsx files with seeded random data.
Run: python gen_fixtures.py
"""
import os
import random
import time

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl")
    raise SystemExit(1)

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")
os.makedirs(OUT_DIR, exist_ok=True)

random.seed(42)

DEPARTMENTS = ["销售", "研发", "市场", "财务", "人事", "运营", "客服", "法务"]
PRODUCTS = ["产品A", "产品B", "产品C", "产品D", "产品E", "产品F", "产品G", "产品H"]
REGIONS = ["华东", "华南", "华北", "西南", "东北", "华中", "西北", "海外"]
NAMES = ["张三", "李四", "王五", "赵六", "钱七", "孙八", "周九", "吴十",
         "郑一", "冯二", "陈三", "楚四", "魏五", "蒋六", "沈七", "韩八"]


def gen_row(cols):
    """Generate a random data row with mixed types."""
    row = []
    for c in range(cols):
        mod = c % 8
        if mod == 0:
            row.append(random.choice(NAMES))
        elif mod == 1:
            row.append(random.choice(DEPARTMENTS))
        elif mod == 2:
            row.append(random.choice(PRODUCTS))
        elif mod == 3:
            row.append(random.choice(REGIONS))
        elif mod == 4:
            row.append(round(random.uniform(1000, 99999), 2))
        elif mod == 5:
            row.append(random.randint(1, 500))
        elif mod == 6:
            row.append(f"2024-{random.randint(1,12):02d}-{random.randint(1,28):02d}")
        else:
            row.append(f"备注_{random.randint(1000,9999)}")
    return row


def gen_header(cols):
    headers = ["姓名", "部门", "产品", "区域", "金额", "数量", "日期", "备注"]
    row = []
    for c in range(cols):
        row.append(headers[c % 8] + (f"_{c//8+1}" if c >= 8 else ""))
    return row


def create_workbook(filename, sheets_config):
    """Create workbook. sheets_config: list of (sheet_name, rows, cols)."""
    print(f"  Creating {filename}...")
    wb = Workbook()
    first = True
    for sheet_name, rows, cols in sheets_config:
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        # Header
        header = gen_header(cols)
        for c, val in enumerate(header, 1):
            ws.cell(row=1, column=c, value=val)

        # Data
        for r in range(2, rows + 2):
            row_data = gen_row(cols)
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=val)

        if r % 10000 == 0:
            print(f"    {sheet_name}: {r}/{rows} rows...")

    path = os.path.join(OUT_DIR, filename)
    t0 = time.time()
    wb.save(path)
    elapsed = time.time() - t0
    size_mb = os.path.getsize(path) / (1024 * 1024)
    print(f"  -> {path} ({size_mb:.1f} MB, saved in {elapsed:.1f}s)")
    return path


def main():
    print("=" * 60)
    print("  Generating benchmark fixtures")
    print("=" * 60)

    # 1. Empty workbook (~6KB)
    print("\n[1/3] Empty workbook")
    wb = Workbook()
    wb.active.title = "Sheet1"
    path = os.path.join(OUT_DIR, "bench_empty.xlsx")
    wb.save(path)
    print(f"  -> {path} ({os.path.getsize(path)} bytes)")

    # 2. Medium (~50MB): 5 sheets x 50000 rows x 30 cols
    # Each cell avg ~10 bytes in xlsx → 5*50000*30*10 ≈ 75MB uncompressed → ~50MB xlsx
    print("\n[2/3] Medium workbook (~50MB target)")
    create_workbook("bench_medium.xlsx", [
        ("销售数据", 50000, 30),
        ("产品明细", 50000, 30),
        ("区域汇总", 50000, 30),
        ("财务报表", 50000, 30),
        ("人员信息", 50000, 30),
    ])

    # 3. Large (~100MB): 5 sheets x 100000 rows x 30 cols
    print("\n[3/3] Large workbook (~100MB target)")
    create_workbook("bench_large.xlsx", [
        ("销售数据", 100000, 30),
        ("产品明细", 100000, 30),
        ("区域汇总", 100000, 30),
        ("财务报表", 100000, 30),
        ("人员信息", 100000, 30),
    ])

    print("\n" + "=" * 60)
    print("  Done! Files in:", OUT_DIR)
    print("=" * 60)


if __name__ == "__main__":
    main()
